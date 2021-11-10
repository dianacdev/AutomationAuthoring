from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.common.exceptions import TimeoutException, ElementClickInterceptedException, NoSuchElementException
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
import tkinter.font as tkFont
from tkinter import *
from tkinter import messagebox, LabelFrame, Frame, filedialog, ttk
from ttkbootstrap import Style, Colors
import os


#AutomationTest.xlsx
def LoadingExcelInfo():
    global sheetindex, questionType, QuestionName,InstructionsToBeEntered, HelpTextToBeEntered,DropTargetToBeEntered,CorrectAnswerCell,Objective,SubObjective,TotalSheets,correctanswerAmount,amountofAnswers,answerIndex,allAnswersClicked,correctAnswerIndex
    #ExcelFileName = ''
    wb = load_workbook(ExcelFileName)
    sheetindex = 0
    sheets = wb.sheetnames
    print("SheetNames: " + str(sheets))
    ws = wb.active
    i = 4
    questionType = ws['C'+str(i)]
    QuestionName = ws['B'+str(i)]
    InstructionsToBeEntered = ws['D'+str(i)]
    HelpTextToBeEntered = ws['H'+str(i)]
    DragOptionToBeEntered = ws['E'+str(i)]
    DropTargetToBeEntered = 'Drop Target Entered'
    CorrectAnswerCell = ws['F'+str(i)]
    CorrectAnswerCellList = ws['F'+str(i)]
    ObjectivesCell = ws['J'+str(i)]
    ObjectivesCell = str(ObjectivesCell.value).split('\n')
    Objective = str(ObjectivesCell[0])
    SubObjective = str(ObjectivesCell[1])
    TotalSheets = len(sheets)
    correctanswerAmount = len(str(CorrectAnswerCellList.value).split('\n'))

    amountofAnswers = len(str(DragOptionToBeEntered.value).split('\n'))
    answerIndex = 0
    allAnswersClicked = 0
    correctAnswerIndex = 0
    #print("Amount of Total Answers"+str(amountofAnswers))

def SheetChecker():
    global i, questionType, QuestionName, InstructionsToBeEntered, HelpTextToBeEntered, DropTargetToBeEntered, CorrectAnswerCell, Objective, SubObjective, correctanswerAmount, amountofAnswers
    print(str(i)+" before")
    i += 1
    if ws.max_row != (i-1):
        print(str(ws.max_row) + ' Max Row Number')
        global CorrectAnswerCellList, ObjectivesCell
        questionType = ws['C'+str(i)]
        QuestionName = ws['B'+str(i)]
        InstructionsToBeEntered = ws['D'+str(i)]
        HelpTextToBeEntered = ws['H'+str(i)]
        DragOptionToBeEntered = ws['E'+str(i)]
        DropTargetToBeEntered = 'Drop Target Entered'
        CorrectAnswerCell = ws['F'+str(i)]
        CorrectAnswerCellList = ws['F'+str(i)]
        ObjectivesCell = ws['J'+str(i)]
        ObjectivesCell = str(ObjectivesCell.value).split('\n')
        Objective = str(ObjectivesCell[0])
        SubObjective = str(ObjectivesCell[1])
        correctanswerAmount = len(str(CorrectAnswerCellList.value).split('\n'))
        amountofAnswers = len(str(DragOptionToBeEntered.value).split('\n'))
        print(str(i)+" after")
        QuestionTypeClicker()
        QuestionCreation()
    else:
        print("Finished cycling!!")
    # print(str(DragOptionToBeEntered.value).split('\n'))
    #print(str(i) + " :This is the i value")

def TimeoutErrorMessage():
    global QIDCell
    QIDCell.fill = PatternFill(fgColor='34B1EB', fill_type='solid')
    QIDCell.value = "Question Failed To Create"
    wb.save(ExcelFileName)

def ClickMoreAnswer():
    global driver
    if amountofAnswers > answerIndex:
        # Create Another Answer
        CreateAnotherAnswer = driver.find_element_by_id(
            "AddAnswer0"+str(answerIndex))
        CreateAnotherAnswer.click()

def FinallyCreateQuestionButton():
    global driver, questionType
    if questionType.value == 'IFrame':
        FinalCreate = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[@id="createQuestion"]/div[3]/div/input'))
        ).click()
    else:
        FinalCreate = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, 'OnClickCreateQuestion'))
        ).click()
    time.sleep(2)

def RetreiveQID():
    global driver, i, QIDCell, QIDName
    QIDName = driver.find_element_by_class_name("modal-body").text
    QIDCell = ws['N'+str(i)]
    if questionType.value == "Multiple Choice":
        QIDCell.fill = PatternFill(fgColor='29FF49', fill_type='solid')
    # elif ws['C'+str(i)].value.lower() == "mcwi":
     #   QIDCell.fill = PatternFill(fgColor='34B1EB', fill_type='solid')
    else:
        QIDCell.fill = PatternFill(fgColor='FF4229', fill_type='solid')

    QIDCell.value = str(QIDName).strip(
        "Question successfully created. Question ID: ")
    wb.save(ExcelFileName)
    print(QIDName)

def QuestionTypeClicker():
    global driver, questionType
    openQuestionDD = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.XPATH, ('//*[@id="questionTypeContainer"]/div[2]/span[1]/span/span[1]'))))
    openQuestionDD.click()
    if questionType.value.lower() == 'multiple choice' or questionType.value.lower() == 'choose' or questionType.value.lower() == 'multiple' or questionType.value.lower() == 'mcwi':
        questionType.value = 'Multiple Choice'
    elif questionType.value.lower() == "dnd" or questionType.value.lower() == "drag and match":
        questionType.value = 'Drag and Match'
    elif questionType.value.lower() == 'applications':
        questionType.value = 'Application'
    elif questionType.value.lower() == 'demo' or questionType.value.lower() == 'iframe' or questionType.value == 'Demo':
        questionType.value = 'IFrame'
    time.sleep(1)
    questionTypeSelected = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="QuestionType-list"]//li[text() = "'+questionType.value+'"]'
                                    ))
    )
    print("Question Type: " + questionType.value)
    questionTypeSelected.click()

def QuestionCreation():
    global questionType, QuestionName, answerIndex, amountofAnswers, DragOptionToBeEntered, DropTargetToBeEntered, InstructionsToBeEntered, HelpTextToBeEntered, i, driver, allAnswersClicked, Objective, SubObjective, correctanswerAmount
    global correctAnswerIndex, CorrectAnswerCell, QIDCell, ErrorMessage
    # Inputting Question Name
    time.sleep(1)
    questionNameEntry = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.ID, 'QuestionName')))
    #questionNameEntry = driver.find_element_by_id("QuestionName")
    questionNameEntry.send_keys(QuestionName.value)
    # Opens Objective Menu
    try:
        objectiveDD = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="ObjectiveId"]'
                                            ))
        )
        objectiveDD.click()
    except:
        print("Object failed to Match any listed under this product")
        ErrorMessage = "Object failed to Match any listed under this product"
        TimeoutErrorMessage()
        ErrorWindowDefault()
        SheetChecker()
        ResetWindow()
        LoginAndOpenQuestionInput()

    time.sleep(1)
    selectionOBJ = Select(driver.find_element_by_xpath(
        "//select[@name='ObjectiveId']"))
    selectionOBJ.select_by_visible_text(Objective.strip(" "))
    # Opens subObjective Drop Down
    subObjectiveDD = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located(
            (By.XPATH, '//*[@id="ddlSubObjective"]'))
    )
    subObjectiveDD.click()
    time.sleep(1)
    selectionSubOBJ = Select(driver.find_element_by_xpath(
        "//select[@name='ddlSubObjective']"))
    try:
        selectionSubOBJ.select_by_visible_text(SubObjective.strip(" "))
    except TimeoutException or ElementClickInterceptedException or NoSuchElementException:
        ErrorMessage = "SubObjective does not match any listed under this product, Click OK to cycle to next Question"
        ErrorWindowDefault()
        TimeoutErrorMessage()
        SheetChecker()
        print("SubObjective does not match any listed under this product")

    time.sleep(1)
    # Instructions
    InstructionFrame = WebDriverWait(driver, 20).until(
        EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="cke_1_contents"]/iframe')))
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
        (By.XPATH, "//body[@class='cke_editable cke_editable_themed cke_contents_ltr cke_show_borders']/p"))).send_keys(InstructionsToBeEntered.value)
    driver.switch_to.default_content()
    # HelpText
    HelpFrame = WebDriverWait(driver, 20).until(
        EC.frame_to_be_available_and_switch_to_it((By.XPATH, '//*[@id="cke_2_contents"]/iframe')))
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
        (By.XPATH, "//body[@class='cke_editable cke_editable_themed cke_contents_ltr cke_show_borders']/p"))).send_keys(HelpTextToBeEntered.value)
    driver.switch_to.default_content()

    if questionType.value == 'Multiple Choice':
        DragOptionToBeEntered = str(DragOptionToBeEntered.value).split('\n')
        CorrectAnswerCell = str(CorrectAnswerCell.value).split('\n')
        while amountofAnswers > answerIndex:
            MultipleChoice1 = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((
                    By.XPATH, "//input[@name='InstructionComponent[0].AnswersBlock.Answers["+str(answerIndex)+"].AnswerText']"))).send_keys(DragOptionToBeEntered[answerIndex])
            if correctanswerAmount != allAnswersClicked:
                if str(CorrectAnswerCell[correctAnswerIndex]) == str(DragOptionToBeEntered[answerIndex]):
                    CorrectAnswerClicked = WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, '//*[@id="InstructionComponent_0__AnswersBlock_Answers_'+str(answerIndex)+'__IsCorrect"]'))
                    ).click()
                    correctAnswerIndex += 1
                    allAnswersClicked += 1
                    time.sleep(1)
            if (int(amountofAnswers)-1) != answerIndex:
                ClickMoreAnswer()
                time.sleep(1)
                #print(str(answerIndex) + "current answer index")
            answerIndex += 1

    FinallyCreateQuestionButton()
    RetreiveQID()
    CloseQIDMenu = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located(
            (By.XPATH, '//*[@id="successfulUpdateModal"]/div/div/div[1]/button/span'))
    )
    CloseQIDMenu.click()
    SheetChecker()

def LoginAndOpenQuestionInput():
    global Username, Password, driver, categoryName, productName, ErrorMessage, ErrorWindow
    if not driver:

        Username = Username_var.get()
        Password = Password_var.get()
        categoryName = " " + Category_var.get()
        productName = Product_var.get()
        PATH = "C:\Program Files (x86)\chromedriver.exe"
        driver = webdriver.Chrome(PATH)
        driver.get("https://author.gmetrix.net")
        # This is opening the webpage^^^
        driver.maximize_window()
        search = driver.find_element_by_id("Email")
        # Your Credentials here
        search.send_keys(Username)
        search = driver.find_element_by_id("Password")
        search.send_keys(Password)
        time.sleep(4)

        signIn = driver.find_element_by_xpath('//*[@id="buttonSignIn"]')
        signIn.click()
        # This is Signing the user in ^^

        questionTab = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.LINK_TEXT, "Questions"))
        )
        questionTab.click()
        createButton = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.LINK_TEXT, "Create"))
        )
        createButton.click()
        # Opening Category Menu
        openCategoryDD = driver.find_element_by_css_selector(
            'span[class="k-input"]')
        openCategoryDD.click()

        # Clicking Category Selected
        categorySelected = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="Category-list"]//li[text() = "'+categoryName+'"]'
                                        ))
        )
        categorySelected.click()
        # Opening Product Drop Down List
        openProductDD = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, ('/html/body/div[2]/form/div[1]/div/div/div[2]/div[2]/span[1]/span/span[1]'))))
        openProductDD.click()
        time.sleep(2)
        # Clicking Product Name
        productSelected = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="Product-list"]//li[text() = "'+productName+'"]'
                                        ))
        )
        productSelected.click()

        time.sleep(2)
        # Opening Question Drop Down List
        # Inputing the Question Type
        QuestionTypeClicker()
        firstCreateButton = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[2]/form/div[2]/input'))
        )
        firstCreateButton.click()
        QuestionCreation()

def DisplayStartWindow():
    global driver, check, PasswordInput, Password_var, UsernameInput, Username_var, Category_var, Product_var, ExcelName_var, root, style,ThemeDD, clicked, mycanvas
    root = Tk()
    style = Style(theme='vaporwave')

    Background = style.colors.bg
    ToggleBGColor = style.colors.secondary
    print(ToggleBGColor)
    ErrorMessage = ""
    Username_var = StringVar()
    Password_var = StringVar()
    Category_var = StringVar()
    Product_var = StringVar()
    ExcelName_var = StringVar()
    root.title("Authoring Automation Tool")
    driver = None
    root.geometry('1000x800')
    root.configure(background=Background)
    LabelFontStyle = tkFont.Font(family="Helvetica", size=12, weight="bold")
    HeaderFontStyle=tkFont.Font(family="Veranda bold", size=20,weight="bold")

    Themes =["minty","flatly","cosmo","litera","lumen","pulse","sandstone","united","yeti","superhero","solar","cyborg","darkly","vaporwave","carnage"]

    # Creating menuFrame
    menuFrame = ttk.Frame(root, style='primary.TFrame')
    menuFrame.pack(pady=50)

    backgroundImage = PhotoImage(file="MicrosoftTeams-image.png")
    backgroundLabel = Label(menuFrame, image=backgroundImage)
    backgroundLabel.place(x=0, y=0, relwidth=1, relheight=1)
    
    entryFrame = ttk.Frame(menuFrame, style='secondary.TFrame')
    entryFrame.pack(padx=60, pady=60)

    productFrame = ttk.Frame(menuFrame, style='secondary.TFrame')
    productFrame.pack(padx=10, pady=10)

    ProjectLabel = ttk.Label(menuFrame, text='Question Automation Tool', font=HeaderFontStyle,
                             style='primary.Inverse.TLabel')
    ProjectLabel.place(x=120, y=10)

    clicked = StringVar()
    clicked.set("Themes")

    drop = OptionMenu(root, clicked, *Themes, command=(changeTheme)).place(x=850, y =10)

    UsernameLabel = ttk.Label(entryFrame, text='Username :',
                              style='secondary.Inverse.TLabel', font=LabelFontStyle)
    UsernameLabel.grid(row=1, column=1, padx=5, pady=5)

    UsernameInput = ttk.Entry(entryFrame, textvariable=Username_var,
                              style='primary.TEntry', width=25).grid(row=1, column=2, padx=5)

    PasswordLabel = ttk.Label(entryFrame, text='Password :', font=LabelFontStyle,
                              style='secondary.Inverse.TLabel').grid(row=2, column=1, padx=10)

    PasswordInput = ttk.Entry(entryFrame, textvariable=Password_var,
                              style='secondary.TEntry', width=25, show='*')
    PasswordInput.grid(row=2, column=2, padx=20)

    check = ttk.Checkbutton(entryFrame, text='Show Password',
                            command=show, style='success.Roundtoggle.Toolbutton')
    check.grid(row=2, column=3, padx=10, pady=5)

    CategoryLabel = ttk.Label(productFrame, text='Enter Category:',
                              style='secondary.Inverse.TLabel', font=LabelFontStyle).grid(row=1, column=1, padx=10, pady=10)
    CategoryInput = ttk.Entry(productFrame, textvariable=Category_var,
                              style='primary.TEntry', width=30).grid(row=1, column=2, padx=10, pady=10)
    ProductLabel = ttk.Label(productFrame, text='Enter Product:',
                             style='secondary.Inverse.TLabel', font=LabelFontStyle).grid(row=2, column=1, padx=10, pady=10)
    ProductInput = ttk.Entry(productFrame, textvariable=Product_var,
                             style='primary.TEntry', width=30).grid(row=2, column=2, padx=10, pady=10)

    BrowseButton = ttk.Button(productFrame, text="Browse Files", command=(browseFiles), style='primary.TButton').grid(row=3, column=2, padx=5, pady=5)

    Start = ttk.Button(root, text="Start", command=(
        LoginAndOpenQuestionInput), style="info.TButton", width=25).place(x=400, y=400)

    # MessageDisplayed = Label(root, text="Status: " + ErrorMessage,
    #                        bg='#1e2124',  fg='#FFF').place(x=10, y=250)
    root.mainloop()

def changeTheme():
    global style, clicked
    style = Style(theme=clicked)

def search_for_file_path ():
    global tempdir
    currdir = os.getcwd()
    tempdir = filedialog.askopenfilename(parent=root, initialdir=currdir, title='Please select a File')
    if len(tempdir) > 0:
        print ("You chose: %s" % tempdir)
    return tempdir

def browseFiles():
    search_for_file_path()
    global ExcelFileName
    ExcelFileName = tempdir
    print ("\nfile_path_variable = ", ExcelFileName)
    LoadingExcelInfo()
      
def show():
    PasswordInput.configure(show='')
    check.configure(command=hide, text='Hide Password')

def hide():
    PasswordInput.configure(show='*')
    check.configure(command=show, text='Show Password')

def ErrorWindowDefault():
    messagebox.showerror(title="Error", message=ErrorMessage)

def ResetWindow():
    root.destroy()
    DisplayStartWindow()

DisplayStartWindow()

# .split('\n')
"""        categorySelected = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="Category-list"]//li[text() = "'+categoryName+'"]'
                                        ))"""







# ~~~~~~~~~~~Made By Diana Cervantes ~~~~~~~~~~~~~ Project: Question Creation Automation


